#Import critical gui functions and libraries for math and Excel manipulation
import tkinter as tk
import openpyxl 
from tkinter import *
from tkinter import ttk
import xlwings as xw
import math
import shutil 

#Update list function pulls from the selected Index of Refraction entry and outputs the critical angle
#Inputs are the selection list, the value chosen, and a new but empty list to assign the index of refraction values 
#value = text value defined in the combobox
#oglist = the original list that supplies values to the constant box (acrylic, glass, silicone, etc.)
def updatedlist(value,oglist):    
    entrytext = tk.StringVar()                                                  #Defines string variable entrytext
    IndexRefVals = [1.49, 1.40, 1.50, 1.33]                                     #Values of Index of Refractions given in combobox. Values must be in same order as the text in the combobox array
    New_Entry = tk.Entry(textvariable = entrytext, width = 10, font = ("30"), state = 'readonly')   #Defines new entry with entrytext where critical angle will appear in readonly state 
    New_Entry.grid(row=5,column =6)                                             #Defines location of new entry                           
    temp = oglist.index(value)                                                  #temp = index of oglist that matches value passed to function 
    ref = IndexRefVals[temp]                                                    #ref = text in IndexRefVals that matches the index of temp 
    CA = math.degrees(math.asin(float(1)/float(ref)))                           #Critical angle math 
    entrytext.set(('{:.4f}'.format(CA) + '\u00B0'))                             #sets value in entry to 4 decimal places and prints the degree symbol
#End of Updatedlist    

#Creation of Critical angle entry
def CA_Entry_create():
    var3 = tk.StringVar()                                                                   #Defines string variable var3
    CA_label = tk.Label(text='Critical Angle of Material',font="30")                        #generates label for critical angle entry
    vlist=["Acrylic","Silicone","Glass","Water"]                                            #defines list of options to feed combobox i.e optical material choices
    CA_combo = ttk.Combobox(width=10, font=("30"))                                          #generates a combobox  
    CA_label.grid(row=4,column=5)                                                           #Defines label position
    CA_combo.grid(row=4,column=6)                                                           #Defines combobox position
    CA_combo['value'] = vlist                                                               #Assigns vlist array to values in combobox
    var3 = CA_combo                                                                         #Defines var3 equal to the choice from the combobox
    testb = tk.Button(text = 'Check', command=lambda: updatedlist(var3.get(), vlist))       #generates button with command for updatedlist, which returns critical angle math
    testb.grid(row=5, column=5)                                                             #defines button position
#end of creation angle entry

#create function for opening new optic profile file and writing values from entries
#q = filename to be created for writing these entries 
def writingfile(q):
    f = open(q,"w+")                                                            #Opens/creates file with name q and assigns that file value f
    for entry in weight_entry:                                                  #for every entry in the weight entry boxes
        f.write('%s\n' %entry.get())                                            #write to f the entries 
    f.close()                                                                   #close the file 
#end function for opening new optic profile file and writing values from entries

#converts file values from strings to floats
#filename = .txt file that you wish to return the string values as floats. 
def read_floats(filename): 
    with open(filename) as f:                                                   #opens/creates the file with the name passed through the function
        return[float(x) for x in f]                                             #return the values from that file as a list of floats for each value in the file
#end conversion file values strings to floats

#Define submit button to generate text file full of constants, export constants to excel, and save new file
def submit():                               
    vers1 = version.get()                                                       #obtains value entered in the file version entry box
    var5 = var4.get()                                                           #obtains value selected from the starting radius entry box
    if(rad.get() == 1):                                                         #if radio button value = 1
        prof = 'Reflector'                                                      #define reflector string
        f2 = ("Reflector Weights" + vers1 + ".txt")                             #define the file name that will store the reflector weights
        writingfile(f2)                                                         #call writingfile for the reflector weights file 
        str2 = 'C:\\Users\\j90742\\OneDrive - 日亜化学工業株式会社\\Documents\\Building a Reflector model_5_6_21.xlsm'      #New string finding reflector excel file path and document name
        generatefile(f2,str2,prof,vers1,var5)                                   #call generatefile based on previous information
    elif(rad.get() == 2):                                                       #if radio button value = 2
        prof = 'Refractor'                                                      #define refractor string
        f2 = ("Refractor Weights" + vers1 + ".txt")                             #define the file name that will store the refractor weights
        writingfile(f2)                                                         #call writingfile for the refractor weights file
        str2 = 'C:\\Users\\j90742\\OneDrive - 日亜化学工業株式会社\\Documents\\Building a Refractor model_5_6_21.xlsm'      #New string finding refractor excel file path and document name 
        generatefile(f2,str2,prof,vers1,var5)                                   #call generate file based on previous information
    else:
         pass                                                                   #if the radio button isn't selected, don't do anything
#End submit button

#Opens Excel file, saves updates, and then closes
def ExcelOnS():
    app = xw.App(visible=False)                                                 #Open the excel application but do not show it on screen
    book = app.books.open(r"Test123.xlsx")                                      #open this specific book called (insert string)
    book.save()                                                                 #Save the book in order to update the formulas with values inserted
    book.close()                                                                #close the book
    app.kill()                                                                  #close the whole app
#End of Opens Excel file, saves updates, and then closes

#function that reads text file, converts to floats, and imports values into excel spreadsheet. Also pulls from new excel file the final angle value
#x = value of full file name for reflector weights text file 
#y = value of string referencing specific excel file path
#z = value that is assigned string value of "reflector" or "refractor"
#t = value passed from file version entry box
#v = Value passed from starting radius entry box
def generatefile(x,y,z,t,v):
    f3 = read_floats(x)                                                         #calls the readfloats file to returns flaots of a created file 
    XLdata1 = openpyxl.load_workbook(y)                                         #opens the workbook from the filepath designated
    XLdata1.sheetnames                                                          #pulls out all of the sheet names
    sheet1 = XLdata1[z + " design sheet"]                                       #defines sheet1 as the specific worksheet in xl 
    for k in range(4,44):                                                       #start in range of the rows on the worksheet
        cellref = sheet1.cell(row= k, column= 11)                               #find sheet cells that match row k in column 11
        cellref.value = f3[k-4]                                                 #assign those cells the values from the read_floats output (in this case, inserts the reflector weights)
    sheet05 = XLdata1["Flux maps " + z + " final"]                              #open new sheet  
    cellref2 = sheet05.cell(row=32, column=18)                                  #locate cell in row 32 and column 18
    cellref2.value = float(v)                                                   #assign cell value to v (in this case, starting radius value)
    XLdata1.save(filename = 'Test123.xlsx')                                     #saves the file as new file, based on string 
    XLdata1.close()                                                             #Closes excel file 
    ExcelOnS()                                                                  #calls ExcelOnS function to update formulas
    XLdata2 = openpyxl.load_workbook('C:\\Users\\j90742\\OneDrive - 日亜化学工業株式会社\Documents\\GitHub\\Python-Early-Codes\\Test123.xlsx', data_only=True)  #opens new workbook that was saved and updated in new path
    XLdata2.sheetnames                                                          #opens all sheet names
    sheet2 = XLdata2[z + " design sheet"]                                       #searches for specific worksheet in xl
    Final_angle_exit = sheet2['J43'].value                                      #Obtains value from cell J43 in that worksheet
    XX = []                                                                     #defines new array XX
    YY = []                                                                     #defines new array YY
    for m in range(50,91):                                                      #in defined range of m
        XX.append(sheet2.cell(row=m,column=11).value)                           #append XX with values fromt he worksheet according to m and column 11
        YY.append(sheet2.cell(row=m,column=12).value)                           #append YY with values fromt he worksheet according to m and column 12
    ExportXY(XX,YY,z,t)                                                         #call exportXY to write files with values
    Final_angle_exit_f = '{:.4f}'.format(Final_angle_exit)                      #formats the final angle exit value 
    FAE_entry.insert(0,(Final_angle_exit_f + '\u00B0'))                         #inserts final angle exit value into entry 
#end of function to read text file, convert floats, import values to spreadsheet, and read final angle value    

#Exports the final CAD file used to create the reflector
#a = array/list that contains values
#b = array/list that contains values
#c = text file name from entrybox that creates the exported file (part 1)
#d = text file name from entrybox that creates the exported file (part 2)
def ExportXY(a,b,c,d):
    xyfile = []                                                                 #created empty array
    f4 = open( c + ' CAD file ' + d + ".txt","w+")                              #open new file with c and d as part of the naming convention
    for w in range (40):                                                        #in the range of all values
        xyfile.append( str(a[w]) + "    " + str(b[w]) + "   " + '0')            #write the values as strings  to xyfile to create a CAD file as a tab deliminated .txt 
    for item in xyfile:                                                         #for each item in xyfile
        f4.write('%s\n' %item)                                                  #write the xyfile to the new file    
    f4.close()                                                                  #close file
    if(c == 'Refractor'):                                                       #Specific case for refractors
        AdjustTIRVals(a,b)                                                      #use the AdjustTIRvals function 
    else:                                                                       #otherwise do nothing
        pass
#End CAD file export

#Function created to export the new values needed to create an ideal TIR refractor geometry
#X = list of values for new refractor curve (x)
#Y = list of values for new refractor curve (y)
def AdjustTIRVals(X,Y):
    xtemp = tk.StringVar()                                                      #temp variable for writing values later
    ytemp = tk.StringVar()                                                      #temp variable for writing values later
    run = X[-1]                                                                 #run defines the last variable in X space
    rise = Y[-1]                                                                #rise defines the last variable in Y space
    x_start = float(X[0])                                                       #turn intial radius value into float and pass
    m = float(rise)/float(run)                                                  #m gives the slope through the lamp center
    y_start = m*x_start                                                         #ystart finds the y point where xstart*slope 
    XAdjust_l = tk.Label(text='X Adjust:',width=10)                             #create Xlabel 
    YAdjust_l = tk.Label(text='Y Adjust:',width=10)                             #create Ylabel
    XAdjust_E = tk.Entry(textvariable = xtemp,width=10, state = 'readonly')     #create Xentry and state = readonly
    YAdjust_E = tk.Entry(text = ytemp,width=10, state = 'readonly')             #create Yentry and state = readonly
    XAdjust_l.grid(row=10,column=5)                                             #define Xlabel position
    XAdjust_E.grid(row=10,column=6)                                             #define X Entry position
    YAdjust_l.grid(row=11,column=5)                                             #define Y label position
    YAdjust_E.grid(row=11,column=6)                                             #define Y entry position
    xtemp.set('{:.4f}'.format(x_start))                                         #set xentry to the x_start
    ytemp.set('{:.4f}'.format(y_start))                                         #set the yentry to the y_start
#End AdjustTIRVals

#clear text function
def clear_text():
    FAE_entry.delete(0,END)                                                     #Deletes whatever is inside of the final angle exit entry
#End fo clear text function

#Move file function which automates moving the CAD file to the Photopia folder
def movefile():
    rad_temp = rad.get()                                                        #Retrieves the value selected from the radio button
    if(rad_temp == 1 or rad_temp == 2):                                         #if the radio button is selected
        phrase = Cad_move.get()                                                 #Call the CAD_move value
        str_temp = 'C:\\Users\\j90742\\OneDrive - 日亜化学工業株式会社\\Documents\\' + phrase + '.stl'      #temp string to find first location of file
        str_temp2 = 'C:\\Users\\Public\\Documents\\LTI Optics\\Photopia'                                  #temp string 2 to find second location of file
        shutil.copy(str_temp,str_temp2)                                                                   #Copy the file from string temp to string temp2                 
    else:
        pass                                                                    #pass if radio button not selected
#End movefile function

#Export function necessary to make .txt file
#q = var7.get() which is the name of the exported file
#l1 = A1L which is the list of measured values
#l2 = A2L which is the lsit of theoretical values
#l3 = A3_5L which is the list of candela differences
#l4 = A3L which is the list of % error values
def ExportP_err(q,l1,l2,l3,l4):
    opentext = q + '.txt'                                                       #defines textfile name from q
    f_err = open(opentext,"w+")                                                 #Opens and creates the textfile named after entry q
    for j in range(19):                                                         #for each value from 1-19
        f_err.write(str(l1[j]) + '   ' + str(l2[j]) + '   ' + str(l3[j]) + '   ' + str(l4[j]) + '\n')   #write each of these lists to the file
    f_err.close()                                                               #close the file
#End of export function

# Percent Error calculation and optional button to export values to table
def submit_p():
    A1L = []                                                                    #List of A1 values (empty)
    A2L = []                                                                    #List of A2 values (empty)
    A3L = []                                                                    #List of A3 values (empty)
    A3_5L = []                                                                  #List of A3_5 values (empty)
    Ideal_dist = [ "4000","3671","3405","3177","2960","2716","2387","1873","979","0","0","0","0","0","0","0","0","0","0"]   #List of ideal distribution values
    for j in range(19):                                                         #for each value from 1-19
        tt = tk.StringVar()                                                     #declare tt as variables
        ideal_entry = tk.Entry(textvariable = tt, font=('calibre',8,'normal'),width=7,state = 'readonly')       #create the ideal distribution entries
        ideal_entry.grid(row=j+1,column=9,pady=10,padx=10)                                                      #define the position of the ideal entry
        tt.set(Ideal_dist[j])                                                   #insert the entries in ideal distribution to the array 
    for err_entry in list1:                                                     #for each entry defined in list1
        A1 = err_entry.get()                                                    #get the value of each error entry
        A1L.append(A1)                                                          #add each value to the new A1 list
    for l in range(19):                                                         #for each value from 1-19
        A2 = Ideal_dist[l]                                                      #get the ideal distribution value
        A2L.append(A2)                                                          #add each value to A2 list
    for i in range(19):                                                         #for each value from 1-19
        A3_5 = (float(A1L[i]) - float(A2L[i]))                                  #subtract actual from theoretical
        if(float(A2L[i]) == 0):                                                 #if the value of theoretical is zero
            A3 = 0                                                              #set A3 equal to zero
        else:                                   
            A3 = abs((A3_5)/float(A2L[i])*100)                                  #else set A3 equal to % error 
        A3L.append(A3)                                                          #add each % error to A3 list
        A3_5L.append(A3_5)                                                      #add difference to A3_5 list
    for k in range(19):                                                         #for each value from 1-19
        Vtext = tk.StringVar()                                                  #define Vtext as string
        Ptext = tk.StringVar()                                                  #define Ptext as string
        V_entry = tk.Entry(textvariable = Vtext,font=('calibre',8,'normal'),width=7, state='readonly')      #define entry for difference in candela
        V_entry.grid(row=k+1,column=10,pady=10, padx=10)                                                    #define position of entry for difference in candela
        Vtext.set(abs(A3_5L[k]))                                                                            #set candela difference entries equal to A3_5 list
        P_entry = tk.Entry(textvariable = Ptext, font=('calibre',8,'normal'),width=7, state='readonly')     #define entry for % error
        P_entry.grid(row=k+1,column=11,pady=10, padx=10)                                                    #define position of entry for % error
        Ptext.set(A3L[k])                                                                                   #set % error entries equal to A3 list
        if(abs(float(A3_5L[k])) <= 100 and float(A2L[k]) != 0):                 #if A2 /=0 and candela difference <= 100
            V_entry.config(readonlybackground='#C6EFCE')                        #assign green background
        elif(float(A3_5L[k]) > 100 and float(A2L[k]) != 0):                     #else if A2 /=0 and candela difference > 100
            V_entry.config(readonlybackground="#00B0F0")                        #assign blue background
        elif(float(A3_5L[k]) < -100 and float(A2L[k]) != 0):                    #else if A2 /=0 and candela difference < -100
            V_entry.config(readonlybackground="#FFC7CE")                        #assign red background
        elif(float(A2L[k])==0):                                                 #else if A2 == 0
            V_entry.config(readonlybackground="Grey")                           #assign grey background
        else:                                                                   #else
            V_entry.config(readonlybackground="White")                          #assign white background
    var7 = tk.StringVar()                                                       #define var7 as string
    ExportP_label = tk.Label(text = 'Filename to export \n (exclude .txt)')     #define label for exportP    
    ExportP = tk.Entry(text = var7)                                             #define entry for ExportP with var7 as value
    ExportP_label.grid(row=15,column=6)                                         #define label position for ExportP
    ExportP.grid(row=16,column=6)                                               #define entry position for ExportP
    button4 = tk.Button(text = 'Export ' + u"\u0025" + ' Error', command=lambda:[ExportP_err(var7.get(),A1L,A2L,A3_5L,A3L)])    #Define button for ExportP and pass variables var7 and all lists
    button4.grid(row=14,column=6)                                                                                               #Define button position for ExportP
#End of Percent Error Calculation

#Create frame, window, title
window = tk.Tk()                                                                #Create window 
frame1 = tk.Frame(master=window,relief=tk.RIDGE)                                #Frame 1 creation
frame1.grid(sticky='news')                                                      #Frame 1 position
title = tk.Label(text='LED Optic Weighting Widget',font="50")                   #create title for master frame
title.grid(row =0,column =5)                                                    #position title in frame
#End of frame, window, and title

#create version for files to write to
version = tk.StringVar()                                                        #define string variable version
version_label = tk.Label(text='File Version',font="30")                         #define label for version entry
version_entry = tk.Entry(textvariable=version,width=2,font=("30"))              #define entry for version entry
version_label.grid(row=0,column=6)                                              #define position of label   
version_entry.grid(row=1,column=6)                                              #define position of entry
#End version to write files to 

#create finale angle entry box
var2 = tk.StringVar()                                                           #define string variable for final angle entry
FAE_label = tk.Label(text='Final Angle of Outer Surface', font="30")            #define label for final angle entry
FAE_entry = tk.Entry(textvariable=var2,width=10,font=("30"))                    #define entry for finale angle entry
FAE_label.grid(row=3,column=5)                                                  #define position of label
FAE_entry.grid(row=3,column=6)                                                  #define position of entry
#end of final entry box

CA_Entry_create()                                                               #call the critical angle entry function

#Create radio button for reflector/refractor choice of design to call the specified spreadsheets
rad = tk.IntVar()                                                                   #create integer value rad
tk.Label(text="Choose optic",padx=20).grid(row=7,column=5)                          #define label for radio button
tk.Radiobutton(text="Reflector",padx=20,variable=rad,value=1).grid(row=8,column=5)  #define radio button Reflector option
tk.Radiobutton(text="Refractor",padx=20,variable=rad,value=2).grid(row=9,column=5)  #define radio button Refractor option
#end radio button creation

#creates entry boxes for angles
weight_entry = []                                                               #define list for weight value entries
for i in range(40):                                                             #for values 1-40
    var1=tk.StringVar()                                                         #create variable string
    if i < 20:                                                                  #for the first 20 values 
        angle1_label = tk.Label(text = "Angle " + str(i+1), font=('calibre',8, 'bold'))     #create label Angle 1, Angle 2, Angle 3....
        angle1_entry = tk.Entry(textvariable = var1, font=('calibre',8,'normal'))           #create entry Angle 1, Angle 2, Angle 3....
        weight_entry.append(angle1_entry)                                       #add each entry to the weight entry list
        angle1_label.grid(row=i,column=0,pady=5)                                #define label position in row i for column 0
        angle1_entry.grid(row=i,column=1,pady=5)                                #define entry position in row i for column 1
    else:                                                                       #for the last 20 entries
        angle1_label = tk.Label(text = "Angle " + str(i+1), font=('calibre',8, 'bold'))     #create label Angle 21, Angle 22, Angle 23....
        angle1_entry = tk.Entry(textvariable = var1, font=('calibre',8,'normal'))           #create entry Angle 21, Angle 22, Angle 23....
        weight_entry.append(angle1_entry)                                       #add each entry to the weight entry list
        angle1_label.grid(row=i-20,column=2,pady=5)                             #define label position for row i-20 for column 2
        angle1_entry.grid(row=i-20,column=3,pady=5)                             #define entry position for row i-20 for column 3  
    angle1_entry.insert(10,"1")                                                 #in all of these, insert the value 1 as a default
#end of angle entry boxes

#Entry for radius adjustment entry box, and autopopulate with a default value of 5
var4 = tk.StringVar()                                                           #define string variable 
Rad_Label = tk.Label(text="Enter Radius value",font=('30'))                     #define label for radius value
Rad_Entry = tk.Entry(text = var4,width = '4')                                   #define entry for radius value
Rad_Label.grid(row = 13,column = 5)                                             #define position for radius value label
Rad_Entry.grid(row = 14,column = 5)                                             #define position for radius value entry
Rad_Entry.insert(0,"5")                                                         #insert default value of 5 
#end Entry for radius adjustment

#call submit button with command
button = tk.Button(text="Click to Create Weighted Optic Profile", command=lambda: [ clear_text(), submit()])    #create button to calculate weighted optic profile
button.grid(row=2,column=5)                                                                                     #define position of button for calculation
#End call submit button 

#Copy and move CAD file into Photopia folder
var6 = tk.StringVar()                                                                       #define string variable
Cad_label = tk.Label(text ="Enter new CAD  \nfilename (exclude .stl)", font=('30'))         #create label for enter CAD filename
Cad_move = tk.Entry(text = var6)                                                            #create entry to input CAD filename
Cad_label.grid(row=7,column=6)                                                              #define position of label
Cad_move.grid(row=8,column=6)                                                               #define position of entry

button2 = tk.Button(text='Click to move STL \n file to Photopia Library', command=lambda:[ movefile()])         #create button for moving .STL file
button2.grid(row=9, column=6)                                                                                   #define position of button
#End copy and move CAD file

#Percent Error Calculation table 
err_above = tk.Label(text = "New Candela")                                      #define label for new candela values
ideal = tk.Label(text = "Ideal")                                                #define label for theoretical candela values
V_above = tk.Label(text = "Error")                                              #define label for candela difference 
P_above = tk.Label(text = u"\u0025" + "Error")                                  #define label for % error
list1 = []                                                                      #define list1 for entries in new candela values
for i in range(19):                                                             #for each value between 1-19
    err_label = tk.Label(text="Angle" + ' ' + str(i*5), font=('calibre',8, 'bold'), width=7)        #define error label for Angle 0, Angle 5, Angle 10...
    err_label.grid(row=i+1,column=7,pady=5,padx=5)                                                  #define label position 
    err_entry = tk.Entry(font=('calibre',8,'normal'),width=7)                                       #define entry for new candela values
    list1.append(err_entry)                                                                         #add entry values to list1
    err_entry.grid(row=i+1,column=8,pady=10,padx=10)                                                #define entry position
err_above.grid(row=0,column=8)                                                  #define label for new candela values position
ideal.grid(row=0,column=9)                                                      #define label for theoretical candela values position
V_above.grid(row=0, column=10)                                                  #define label for candela difference position
P_above.grid(row=0,column=11)                                                   #define label for % error position
button3 = tk.Button(text = 'Percent Error Calc', command=lambda:[submit_p()])                       #define button for calculating % error calculations
button3.grid(row=13,column=6)                                                                       #define button position
#End of Percent Error Calculation Table

window.mainloop()